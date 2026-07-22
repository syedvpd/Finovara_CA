"""PDF and spreadsheet generation.

Documents are produced in memory and returned as bytes; storing them is the
caller's concern. Every amount is formatted from ``Decimal`` — never a float.
"""

from __future__ import annotations

import csv
import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    KeepTogether,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.core.logging import get_logger

logger = get_logger(__name__)

BRAND = colors.HexColor("#087F5B")
INK = colors.HexColor("#102A43")
MUTED = colors.HexColor("#52606D")
RULE = colors.HexColor("#E4E7EB")

_styles = getSampleStyleSheet()
TITLE = ParagraphStyle("FvTitle", parent=_styles["Heading1"], fontSize=18, textColor=INK, spaceAfter=2)
SUBTITLE = ParagraphStyle("FvSub", parent=_styles["Normal"], fontSize=9, textColor=MUTED, spaceAfter=10)
H2 = ParagraphStyle("FvH2", parent=_styles["Heading2"], fontSize=12, textColor=INK, spaceBefore=12, spaceAfter=6)
BODY = ParagraphStyle("FvBody", parent=_styles["Normal"], fontSize=9, textColor=INK, leading=13)
RIGHT = ParagraphStyle("FvRight", parent=BODY, alignment=TA_RIGHT)
FOOT = ParagraphStyle("FvFoot", parent=_styles["Normal"], fontSize=7.5, textColor=MUTED, alignment=TA_CENTER)


def rupees(amount: Decimal | str | int | None) -> str:
    """Indian-format currency, e.g. ₹12,34,567.89."""
    value = Decimal(str(amount or 0))
    negative = value < 0
    whole, _, frac = f"{abs(value):.2f}".partition(".")

    # Indian grouping: last three digits, then pairs.
    if len(whole) > 3:
        head, tail = whole[:-3], whole[-3:]
        parts = []
        while len(head) > 2:
            parts.insert(0, head[-2:])
            head = head[:-2]
        if head:
            parts.insert(0, head)
        whole = ",".join(parts + [tail])

    return f"{'-' if negative else ''}Rs. {whole}.{frac}"


def _fmt_date(value: date | datetime | str | None) -> str:
    if value is None:
        return "-"
    if isinstance(value, str):
        return value
    return value.strftime("%d %b %Y")


def _header(firm: dict[str, Any], title: str, subtitle: str = "") -> list[Any]:
    name = firm.get("legal_name") or "Finovara Chartered Accountants LLP"
    address = firm.get("registered_address") or ""
    identifiers = " · ".join(
        part for part in (
            f"GSTIN: {firm['gstin']}" if firm.get("gstin") else "",
            f"PAN: {firm['pan']}" if firm.get("pan") else "",
        ) if part
    )

    table = Table(
        [[
            Paragraph(f"<b><font size=13 color='#087F5B'>{name}</font></b><br/>"
                      f"<font size=8 color='#52606D'>{address}<br/>{identifiers}</font>", BODY),
            Paragraph(f"<b>{title}</b><br/><font size=8 color='#52606D'>{subtitle}</font>", RIGHT),
        ]],
        colWidths=[105 * mm, 65 * mm],
    )
    table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LINEBELOW", (0, 0), (-1, -1), 1, BRAND),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    return [table, Spacer(1, 8)]


def _kv_block(rows: list[tuple[str, str]], width: float = 85 * mm) -> Table:
    table = Table([[Paragraph(f"<font color='#52606D'>{k}</font>", BODY), Paragraph(v, BODY)] for k, v in rows],
                  colWidths=[width * 0.45, width * 0.55])
    table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
    ]))
    return table


def _data_table(header: list[str], rows: list[list[str]], widths: list[float], right_align: set[int]) -> Table:
    body = [[Paragraph(f"<b>{h}</b>", BODY) for h in header]]
    for row in rows:
        body.append([
            Paragraph(cell, RIGHT if index in right_align else BODY)
            for index, cell in enumerate(row)
        ])

    table = Table(body, colWidths=widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F0F4F8")),
        ("LINEBELOW", (0, 0), (-1, 0), 0.75, BRAND),
        ("GRID", (0, 0), (-1, -1), 0.25, RULE),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFBFC")]),
    ]))
    return table


def _build(elements: list[Any], title: str) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=20 * mm, rightMargin=20 * mm, topMargin=16 * mm, bottomMargin=16 * mm,
        title=title, author="Finovara Chartered Accountants LLP",
    )

    def page_furniture(canvas: Any, _doc: Any) -> None:
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(MUTED)
        canvas.drawCentredString(
            A4[0] / 2, 10 * mm,
            f"Generated {datetime.now().strftime('%d %b %Y %H:%M')}  ·  Page {canvas.getPageNumber()}",
        )
        canvas.restoreState()

    doc.build(elements, onFirstPage=page_furniture, onLaterPages=page_furniture)
    return buffer.getvalue()


# --- Documents ---------------------------------------------------------------


def invoice_pdf(invoice: dict[str, Any], items: list[dict[str, Any]], client: dict[str, Any],
                firm: dict[str, Any]) -> bytes:
    """Tax invoice with a per-line GST breakdown."""
    elements = _header(firm, "TAX INVOICE", invoice.get("invoice_number", ""))

    meta = Table([[
        _kv_block([
            ("Billed to", f"<b>{client.get('legal_name') or client.get('client_code', '')}</b>"),
            ("Address", client.get("registered_address", "-")),
            ("GSTIN", client.get("gst_number") or "-"),
            ("PAN", client.get("pan") or "-"),
        ]),
        _kv_block([
            ("Invoice no.", f"<b>{invoice.get('invoice_number', '')}</b>"),
            ("Issue date", _fmt_date(invoice.get("issue_date"))),
            ("Due date", _fmt_date(invoice.get("due_date"))),
            ("Status", str(invoice.get("status", "")).replace("_", " ").title()),
        ]),
    ]], colWidths=[90 * mm, 80 * mm])
    meta.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    elements += [meta, Spacer(1, 10)]

    rows = []
    for index, item in enumerate(items, start=1):
        rows.append([
            str(index),
            item.get("description", ""),
            str(item.get("quantity", 1)),
            rupees(item.get("unit_price")),
            f"{Decimal(str(item.get('tax_rate_percent', 0))):.2f}%",
            rupees(item.get("tax_amount")),
            rupees(item.get("total_amount")),
        ])

    elements.append(_data_table(
        ["#", "Description", "Qty", "Rate", "GST", "Tax", "Amount"],
        rows,
        [10 * mm, 58 * mm, 12 * mm, 24 * mm, 16 * mm, 24 * mm, 26 * mm],
        right_align={2, 3, 4, 5, 6},
    ))

    totals = [
        ("Subtotal", rupees(invoice.get("subtotal"))),
        ("Discount", f"({rupees(invoice.get('discount_amount'))})"),
        ("GST", rupees(invoice.get("tax_amount"))),
        ("Total", f"<b>{rupees(invoice.get('total_amount'))}</b>"),
        ("Paid", rupees(invoice.get("paid_amount"))),
        ("Amount due", f"<b>{rupees(invoice.get('outstanding_amount'))}</b>"),
    ]
    summary = Table(
        [[Paragraph(f"<font color='#52606D'>{k}</font>", RIGHT), Paragraph(v, RIGHT)] for k, v in totals],
        colWidths=[38 * mm, 32 * mm], hAlign="RIGHT",
    )
    summary.setStyle(TableStyle([
        ("LINEABOVE", (0, 3), (-1, 3), 0.5, RULE),
        ("LINEABOVE", (0, 5), (-1, 5), 0.75, BRAND),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements += [Spacer(1, 8), summary]

    if invoice.get("notes"):
        elements += [Spacer(1, 10), Paragraph(f"<b>Notes</b><br/>{invoice['notes']}", BODY)]

    elements += [
        Spacer(1, 14),
        Paragraph("This is a computer-generated invoice and does not require a signature.", FOOT),
    ]
    return _build(elements, f"Invoice {invoice.get('invoice_number', '')}")


def payslip_pdf(record: dict[str, Any], profile: dict[str, Any], run: dict[str, Any],
                firm: dict[str, Any]) -> bytes:
    """Employee payslip: earnings, statutory deductions, net pay."""
    period = f"{date(run['period_year'], run['period_month'], 1):%B %Y}"
    elements = _header(firm, "PAYSLIP", period)

    elements += [Table([[
        _kv_block([
            ("Employee", f"<b>{profile.get('employee_name', '')}</b>"),
            ("Employee code", profile.get("employee_code", "-")),
            ("PAN", profile.get("pan") or "-"),
            ("UAN", profile.get("uan") or "-"),
        ]),
        _kv_block([
            ("Pay period", period),
            ("Paid days", str(record.get("attendance_days", "-"))),
            ("Loss of pay", str(record.get("lwp_days", "0"))),
            ("Bank", f"****{str(profile.get('bank_account_no', ''))[-4:]}"),
        ]),
    ]], colWidths=[90 * mm, 80 * mm]), Spacer(1, 10)]

    gross = Decimal(str(record.get("gross_salary", 0)))
    basic = Decimal(str(record.get("basic_salary_earned", 0)))
    earnings = [["Basic salary", rupees(basic)], ["Allowances", rupees(gross - basic)]]
    deductions = [
        ["Provident fund", rupees(record.get("pf_deduction"))],
        ["ESI", rupees(record.get("esi_deduction"))],
        ["Professional tax", rupees(record.get("pt_deduction"))],
        ["TDS", rupees(record.get("tds_deduction"))],
        ["Other", rupees(record.get("other_deductions"))],
    ]

    side_by_side = Table([[
        _data_table(["Earnings", "Amount"], earnings, [50 * mm, 32 * mm], {1}),
        _data_table(["Deductions", "Amount"], deductions, [50 * mm, 32 * mm], {1}),
    ]], colWidths=[85 * mm, 85 * mm])
    side_by_side.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    elements.append(side_by_side)

    net = Table([
        [Paragraph("<b>Gross earnings</b>", BODY), Paragraph(f"<b>{rupees(gross)}</b>", RIGHT)],
        [Paragraph("Total deductions", BODY), Paragraph(rupees(
            sum(Decimal(str(record.get(k, 0))) for k in
                ("pf_deduction", "esi_deduction", "pt_deduction", "tds_deduction", "other_deductions"))
        ), RIGHT)],
        [Paragraph("<b>Net pay</b>", BODY), Paragraph(f"<b>{rupees(record.get('net_salary'))}</b>", RIGHT)],
    ], colWidths=[120 * mm, 50 * mm])
    net.setStyle(TableStyle([
        ("LINEABOVE", (0, 2), (-1, 2), 0.75, BRAND),
        ("BACKGROUND", (0, 2), (-1, 2), colors.HexColor("#EAF4F0")),
        ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements += [Spacer(1, 10), net, Spacer(1, 14),
                 Paragraph("Computer-generated payslip; no signature required.", FOOT)]
    return _build(elements, f"Payslip {profile.get('employee_code', '')} {period}")


def tabular_report_pdf(title: str, subtitle: str, header: list[str], rows: list[list[str]],
                       firm: dict[str, Any], widths: list[float] | None = None,
                       right_align: set[int] | None = None, summary: list[tuple[str, str]] | None = None) -> bytes:
    """Generic tabular report used by compliance, trial balance, ageing, GST."""
    elements = _header(firm, title.upper(), subtitle)

    if not rows:
        elements.append(Paragraph("<i>No records for the selected criteria.</i>", BODY))
    else:
        available = 170 * mm
        column_widths = widths or [available / len(header)] * len(header)
        elements.append(_data_table(header, rows, column_widths, right_align or set()))

    if summary:
        block = Table(
            [[Paragraph(f"<font color='#52606D'>{k}</font>", RIGHT), Paragraph(f"<b>{v}</b>", RIGHT)]
             for k, v in summary],
            colWidths=[45 * mm, 35 * mm], hAlign="RIGHT",
        )
        block.setStyle(TableStyle([("LINEABOVE", (0, 0), (-1, 0), 0.75, BRAND),
                                   ("TOPPADDING", (0, 0), (-1, -1), 4)]))
        elements += [Spacer(1, 8), block]

    return _build(elements, title)


def audit_report_pdf(audit: dict[str, Any], observations: list[dict[str, Any]], client: dict[str, Any],
                     firm: dict[str, Any]) -> bytes:
    """Audit report with observations, risk ratings and management responses."""
    elements = _header(
        firm, "AUDIT REPORT",
        f"{str(audit.get('audit_type', '')).title()} audit · {client.get('legal_name') or client.get('client_code', '')}",
    )
    elements += [_kv_block([
        ("Period", f"{_fmt_date(audit.get('start_date'))} to {_fmt_date(audit.get('end_date'))}"),
        ("Status", str(audit.get("status", "")).replace("_", " ").title()),
        ("Overall risk", str(audit.get("risk_rating", "")).title()),
        ("Observations", str(len(observations))),
    ], width=170 * mm), Spacer(1, 6)]

    if not observations:
        elements.append(Paragraph("<i>No observations were raised.</i>", BODY))

    for index, observation in enumerate(observations, start=1):
        responses = observation.get("responses") or []
        reply = responses[0].get("response_text") if responses else "<i>No management response recorded.</i>"
        risk = str(observation.get("risk_level", "")).title()
        elements.append(KeepTogether([
            Paragraph(f"<b>{index}. {observation.get('title', '')}</b>  "
                      f"<font size=8 color='#52606D'>[{risk} risk · "
                      f"{str(observation.get('status', '')).replace('_', ' ').title()}]</font>", H2),
            Paragraph(observation.get("description", ""), BODY),
            Spacer(1, 4),
            Paragraph(f"<font color='#52606D'><b>Management response:</b></font> {reply}", BODY),
            Spacer(1, 8),
        ]))

    return _build(elements, f"Audit report {audit.get('id', '')}")


# --- Spreadsheet / CSV -------------------------------------------------------


def rows_to_csv(header: list[str], rows: list[list[Any]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(header)
    writer.writerows(rows)
    return buffer.getvalue().encode("utf-8-sig")


def rows_to_xlsx(sheet_name: str, header: list[str], rows: list[list[Any]]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name[:31] or "Report"

    sheet.append(header)
    fill = PatternFill("solid", fgColor="F0F4F8")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="102A43")
        cell.fill = fill
        cell.alignment = Alignment(vertical="top")

    for row in rows:
        sheet.append([float(v) if isinstance(v, Decimal) else v for v in row])

    for index, column in enumerate(header, start=1):
        longest = max([len(str(column))] + [len(str(r[index - 1])) for r in rows[:500]] or [10])
        sheet.column_dimensions[get_column_letter(index)].width = min(max(longest + 2, 10), 50)

    sheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
